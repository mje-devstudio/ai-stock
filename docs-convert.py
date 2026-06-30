import os
import re
import openpyxl

def extract_filename(sheet_title):
    """시트 제목에서 괄호() 안의 문자열을 찾아 파일명으로 사용합니다."""
    match = re.search(r'\(([^)]+)\)', sheet_title)
    if match and match.group(1).strip():
        filename = match.group(1).strip()
    else:
        filename = sheet_title.strip()
    
    # 파일명 금지 문자 정제
    filename = re.sub(r'[\/:*?"<>|]', '_', filename)
    return filename

def excel_to_markdown():
    excel_file = "키움 REST API 문서.xlsx"
    if not os.path.exists(excel_file):
        print(f"오류: '{excel_file}' 파일이 존재하지 않습니다.")
        return

    output_dir = "tr_docs"
    os.makedirs(output_dir, exist_ok=True)

    print("엑셀 파일 로드 중...")
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        filename = extract_filename(sheet_name)
        md_filepath = os.path.join(output_dir, f"{filename}.md")
        
        markdown_lines = []
        markdown_lines.append(f"# {sheet_name}\n")
        
        is_first_row = True
        max_columns = 0
        
        for row in sheet.iter_rows(values_only=True):
            # 모든 셀이 비어있는 행은 패스
            if all(cell is None for cell in row):
                continue
                
            processed_row = []
            for cell in row:
                if cell is None:
                    processed_row.append("")
                else:
                    # 표 내부 줄바꿈 처리 및 파이프 기호 이스케이프
                    val = str(cell).replace('\n', '<br>').replace('|', '\|')
                    processed_row.append(val)
            
            # 마크다운 테이블 행 생성
            row_str = "| " + " | ".join(processed_row) + " |"
            markdown_lines.append(row_str)
            
            # [핵심] 첫 번째 행(헤더)을 작성한 직후, 마크다운 표 구분선(---)을 강제로 삽입
            if is_first_row:
                max_columns = len(processed_row)
                separator = "| " + " | ".join(["---"] * max_columns) + " |"
                markdown_lines.append(separator)
                is_first_row = False
        
        # 파일 저장
        with open(md_filepath, "w", encoding="utf-8") as f:
            f.write("\n".join(markdown_lines))
            
        print(f"변환 완료: {md_filepath}")

    print("\n수정된 본으로 모든 시트 변환이 완료되었습니다! 다시 확인해 보세요.")

if __name__ == "__main__":
    excel_to_markdown()